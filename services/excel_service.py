import os
import zipfile

import openpyxl
import pandas as pd
import streamlit as st


@st.cache_data
def extract_all_images(file_path, output_dir):
    with zipfile.ZipFile(file_path, "r") as archive:
        os.makedirs(output_dir, exist_ok=True)

        image_files = [f for f in archive.namelist() if f.startswith("xl/media/")]
        images_info = []

        for i, image_file in enumerate(image_files, start=1):
            image_name = f"image_{i}.jpeg"
            image_path = os.path.join(output_dir, image_name)
            with open(image_path, "wb") as img_file:
                img_file.write(archive.read(image_file))
            images_info.append({"image_name": image_name, "image_path": image_path})

        return images_info


@st.cache_data
def rename_images_based_on_sheet(file_path, output_dir):
    try:
        excel_data = pd.read_excel(file_path, sheet_name=0)
    except Exception as e:
        st.error(f"An error occurred while reading the Excel file: {e}")
        return

    extract_all_images(file_path, output_dir)

    for idx, row in excel_data.iterrows():
        name = row.get("Name")
        if pd.notna(name):
            old_image_path = os.path.join(output_dir, f"image_{idx + 1}.jpeg")
            new_image_path = os.path.join(output_dir, f"{name}.jpeg")
            if os.path.exists(old_image_path):
                os.rename(old_image_path, new_image_path)


def extract_links(uploaded_file, links_column="links"):
    if uploaded_file.name.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        extracted_links = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            if links_column not in headers:
                continue

            links_col_idx = headers.index(links_column) + 1
            for row in sheet.iter_rows(min_row=2):
                link_cell = row[links_col_idx - 1]
                extracted_link = link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value
                if extracted_link:
                    extracted_links.append(extracted_link)

        return extracted_links

    raise ValueError("The uploaded file is not an Excel file.")
