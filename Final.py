import streamlit as st
from app_entry import run_app

run_app()
st.stop()

import shutil
import zipfile
import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from zipfile import ZipFile
from PIL import Image, UnidentifiedImageError
import re
from transformers import pipeline
from transformers.modeling_utils import PreTrainedModel
from collections import defaultdict
import pypdfium2 as pdfium
import os

@st.cache_data
def convert_drive_link(link):
    # Try to match the link with /d/ pattern
    match_d = re.search(r'/d/([^/]+)', link)
    if match_d:
        file_id = match_d.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    
    # Try to match the link with id= pattern or general pattern for Google Drive ID
    match_id = re.search(r'id=([^&]+)|([-\\w]{25,})', link)
    if match_id:
        file_id = match_id.group(1) if match_id.group(1) else match_id.group(2)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    
    # Return the original link if no patterns matched
    return link

@st.cache_data
def download_image(url):
    response = requests.get(url)
    if response.status_code == 200:
        return response.content
    return None

@st.cache_data
def resize_image(image_content, size=(1024, 1024), aspect_ratio_threshold=2):
    try:
        image = Image.open(BytesIO(image_content))
        if image.mode not in ['RGB', 'RGBA']:
            image = image.convert('RGB')

        if image.mode == 'RGBA':
            image = image.convert('RGB')
        
        original_width, original_height = image.size
        aspect_ratio = original_width / original_height
        inverse_threshold = 1 / aspect_ratio_threshold

        if aspect_ratio < inverse_threshold:
            new_height = int(original_width / inverse_threshold)
            crop_top = (original_height - new_height) // 2
            image = image.crop((0, crop_top, original_width, crop_top + new_height))
        elif aspect_ratio > aspect_ratio_threshold:
            new_width = int(original_height * aspect_ratio_threshold)
            crop_left = (original_width - new_width) // 2
            image = image.crop((crop_left, 0, crop_left + new_width, original_height))
        
        image = image.resize(size)
        img_byte_arr = BytesIO()
        image.save(img_byte_arr, format='JPEG')
        return img_byte_arr.getvalue()
    except UnidentifiedImageError:
        return None

if not hasattr(PreTrainedModel, "all_tied_weights_keys"):
    # Some remote-code models do not define this attribute, but newer
    # transformers versions expect it during model load.
    PreTrainedModel.all_tied_weights_keys = {}


@st.cache_resource
def get_rmbg_pipeline():
    return pipeline(
        "image-segmentation",
        model="briaai/RMBG-1.4",
        trust_remote_code=True,
        model_kwargs={"low_cpu_mem_usage": False},
    )


@st.cache_data
def remove_background(image_content):
    try:
        image = Image.open(BytesIO(image_content))
        output_img = get_rmbg_pipeline()(image)
        img_byte_arr = BytesIO()
        output_img.save(img_byte_arr, format='PNG')
        return img_byte_arr.getvalue()
    except UnidentifiedImageError:
        return None

@st.cache_data
def combine_with_background(foreground_content, background_content, resize_foreground=False):
    try:
        foreground = Image.open(BytesIO(foreground_content)).convert("RGBA")
        background = Image.open(BytesIO(background_content)).convert("RGBA")
        background = background.resize((1024, 1024))

        if resize_foreground:
            fg_area = foreground.width * foreground.height
            bg_area = background.width * background.height
            scale_factor = (0.8 * bg_area / fg_area) ** 0.5

            new_width = int(foreground.width * scale_factor)
            new_height = int(foreground.height * scale_factor)

            foreground = foreground.resize((new_width, new_height))
            dimensions = (new_width, new_height)
        else:
            dimensions = (foreground.width, foreground.height)

        fg_width, fg_height = foreground.size
        bg_width, bg_height = background.size
        position = ((bg_width - fg_width) // 2, (bg_height - fg_height) // 2)

        combined = background.copy()
        combined.paste(foreground, position, foreground)
        img_byte_arr = BytesIO()
        combined.save(img_byte_arr, format='PNG')
        return img_byte_arr.getvalue(), dimensions
    except UnidentifiedImageError:
        return None, None

@st.cache_data
def download_all_images_as_zip(images_info, remove_bg=False, add_bg=False, bg_image=None, resize_foreground=False, threshold=2):
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zf:
        for name, url_or_file in images_info:
            if isinstance(url_or_file, str):
                url = convert_drive_link(url_or_file)
                image_content = download_image(url)
            elif isinstance(url_or_file, bytes):
                image_content = url_or_file
            else:
                image_content = url_or_file.read()

            if image_content:
                if remove_bg:
                    processed_image = remove_background(image_content)
                    ext = 'png'
                else:
                    size = (1290, 789) if "banner" in name.lower() else (1024, 1024)
                    processed_image = resize_image(image_content, size=size, aspect_ratio_threshold=threshold)
                    ext = "png"

                if add_bg and bg_image:
                    processed_image, dimensions = combine_with_background(processed_image, bg_image, resize_foreground=resize_foreground)
                    ext = 'png'
                
                # Apply flipping based on user selection
                if flip_horizontal or flip_vertical:
                    image = Image.open(BytesIO(image_content))
                    processed_image = flip_image(image, flip_horizontal, flip_vertical)
                    img_byte_arr = BytesIO()
                    processed_image.save(img_byte_arr, format='PNG')
                    processed_image = img_byte_arr.getvalue()

                if processed_image:
                    zf.writestr(f"{name.rsplit('.', 1)[0]}.{ext}", processed_image)
    zip_buffer.seek(0)
    return zip_buffer

@st.cache_data
def extract_all_images(file_path, output_dir):
    with zipfile.ZipFile(file_path, 'r') as archive:
        os.makedirs(output_dir, exist_ok=True)
        
        image_files = [f for f in archive.namelist() if f.startswith('xl/media/')]
        images_info = []
        
        for i, image_file in enumerate(image_files, start=1):
            image_name = f"image_{i}.jpeg"
            image_path = os.path.join(output_dir, image_name)
            with open(image_path, 'wb') as img_file:
                img_file.write(archive.read(image_file))
            images_info.append({'image_name': image_name, 'image_path': image_path})
            print(f"Extracted {image_name}")
        
        return images_info

@st.cache_data
def rename_images_based_on_sheet(file_path, output_dir):
    try:
        excel_data = pd.read_excel(file_path, sheet_name=0)
    except Exception as e:
        st.error(f"An error occurred while reading the Excel file: {e}")
        return

    extracted_images = extract_all_images(file_path, output_dir)
    
    for idx, row in excel_data.iterrows():
        name = row.get('Name')
        if pd.notna(name):
            old_image_path = os.path.join(output_dir, f"image_{idx + 1}.jpeg")
            new_image_path = os.path.join(output_dir, f"{name}.jpeg")
            if os.path.exists(old_image_path):
                os.rename(old_image_path, new_image_path)
                print(f"Renamed {old_image_path} to {new_image_path}")

#-------------------------- Convert PDF to images-----------------------------
@st.cache_data
def convert_pdf_to_images(pdf_content):
    try:
        images = convert_from_bytes(pdf_content)
        return images
    except Exception as e:
        st.error(f"Error converting PDF to images: {e}")
        return []
#-------------------------- Convert PDF to images-----------------------------
                
def flip_image(image, flip_horizontal=False, flip_vertical=False):
    if flip_horizontal:
        image = ImageOps.mirror(image)
    if flip_vertical:
        image = ImageOps.flip(image)
    return image
#---------------------extract links from hypridlink--------------------------------
import openpyxl

def extract_links(uploaded_file, links_column='links'):
    # Ensure the file is an Excel file
    if uploaded_file.name.endswith('.xlsx'):
        # Load workbook using openpyxl
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        extracted_links = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Find the index of the 'links' column (assuming first row contains headers)
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            if links_column not in headers:
                continue  # Skip this sheet if the 'links' column is not found
            
            links_col_idx = headers.index(links_column) + 1  # openpyxl uses 1-based index for columns
            
            # Iterate over the rows, starting from the second row (skipping the header)
            for row in sheet.iter_rows(min_row=2):
                link_cell = row[links_col_idx - 1]  # Get the cell for the 'links' column

                # Check if the cell contains a hyperlink using openpyxl
                if link_cell.hyperlink:
                    extracted_link = link_cell.hyperlink.target  # Extract the hyperlink URL
                else:
                    extracted_link = link_cell.value  # Use the plain value in the cell if no hyperlink

                # Append the link (either hyperlink target or plain value)
                if extracted_link:
                    extracted_links.append(extracted_link)

        return extracted_links
    else:
        raise ValueError("The uploaded file is not an Excel file.")


#-----------------------------extract links from hypridlink------------------------


# Streamlit app
st.set_page_config(page_title="PhotoMaster", page_icon="🖼️")
st.title("🖼️ PhotoMaster")

# Page layout
col1, col2 = st.columns([2, 1])
threshold = 2.0

with col1:
    
    # Specify the file types you want to accept
    uploaded_files = st.file_uploader("", type=["xlsx", "csv", "jpg", "jpeg", "png", "jfif", "avif", "webp", "heic","NEF","ARW","tiff", "pdf"], accept_multiple_files=True)

    # Input field for Google Drive folder link
    folder_link = st.text_input("Enter Google Drive Link for (**Larger Files**)")

with col2:
    st.markdown("")
    remove_bg = st.checkbox("Remove background")
    add_bg = st.checkbox("Add background")
    resize_fg = st.checkbox("Resize")
    if resize_fg:
        udvanced = st.checkbox("💎Advanced Resize Options")
        if udvanced:
            threshold = st.slider("Aspect Ratio Threshold", 1.0, 2.5, 1.5)
    st.checkbox("👊Compress and Convert Format")
    st.button("Submit")

images_info = []

if  uploaded_files:
    if len(uploaded_files) == 1 and uploaded_files[0].name.endswith(('.xlsx', '.csv')):
        st.write("Select the type of images in the Excel file:")
        images_type = st.radio("Images are:", ["Links of images", "Embedded in Excel file"])
        file_type = 'excel'
    elif all(file.type.startswith('image/') for file in uploaded_files):
        file_type = 'images'
    elif any(file.type == 'application/pdf' for file in uploaded_files):
        file_type = 'pdf'
    else:
        file_type = 'mixed'

    if file_type == 'mixed':
        st.error("You should work with one type of file: either an Excel file, images, or a PDF.")
    else:
        if file_type == 'excel' and images_type == "Links of images":
            uploaded_file = uploaded_files[0]
            if uploaded_file.name.endswith('.xlsx'):
                xl = pd.ExcelFile(uploaded_file)
                
                for sheet_name in xl.sheet_names:
                    st.write(f"Processing sheet: {sheet_name}")  # Debugging print
                    df = xl.parse(sheet_name)
                    if 'links' in df.columns and ('name' in df.columns):
                        df.dropna(subset=['links'], inplace=True)
                        name_count = defaultdict(int)
                        empty_count = 0
                        unique_images_info = []

                        #----function extract links from hyprid links in excel
                        links = extract_links(uploaded_file, links_column='links')
                        #-----------------------------------------------------------
                        #for name, link in zip(df['name'], df['links']):

                        for name, link in zip(df['name'], links):
                            if pd.isna(name) or name.strip() == "":
                                empty_name = f"empty_{empty_count}" if empty_count > 0 else "empty"
                                name = empty_name
                                empty_count += 1
                            if name_count[name] > 0:
                                unique_name = f"{name}_{name_count[name]}"
                            else:
                                unique_name = name
                            unique_images_info.append((unique_name, link))
                            name_count[name] += 1
                        
                        images_info.extend(unique_images_info)
                        if empty_count > 0:
                            st.warning(f"Number of empty cells in 'name' column: {empty_count}")
                    else:
                        st.error(f"The sheet '{sheet_name}' must contain 'links' and 'name' columns.")
            else:
                df = pd.read_csv(uploaded_file)
                if 'links' in df.columns and ('name' in df.columns or 'names' in df.columns):
                    df.dropna(subset=['links'], inplace=True)
                    name_count = defaultdict(int)
                    empty_count = 0
                    unique_images_info = []
                    for name, link in zip(df['name'], df['links']):
                        if pd.isna(name) or name.strip() == "":
                            empty_name = f"empty_{empty_count}" if empty_count > 0 else "empty"
                            name = empty_name
                            empty_count += 1
                        if name_count[name] > 0:
                            unique_name = f"{name}_{name_count[name]}"
                        else:
                            unique_name = name
                        unique_images_info.append((unique_name, link))
                        name_count[name] += 1
                    images_info.extend(unique_images_info)
                    if empty_count > 0:
                        st.warning(f"Number of empty cells in 'name' column: {empty_count}")
                else:
                    st.error("The uploaded file must contain 'links' and 'name' columns.")
        
        elif file_type == 'excel' and images_type == "Embedded in Excel file":
            temp_dir = "temp"
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            
            file_path = os.path.join(temp_dir, uploaded_files[0].name)
            with open(file_path, "wb") as f:
                f.write(uploaded_files[0].getbuffer())
            
            if uploaded_files[0].name.endswith((".xlsx", ".xls")):
                df = pd.read_excel(file_path, sheet_name=0)
            elif uploaded_files[0].name.endswith(".csv"):
                df = pd.read_csv(file_path)
            
            output_dir = os.path.join(temp_dir, "extracted_images")
            if os.path.exists(output_dir):
                shutil.rmtree(output_dir)
            
            os.makedirs(output_dir, exist_ok=True)
            rename_images_based_on_sheet(file_path, output_dir)
            images_info = [(image, open(os.path.join(output_dir, image), "rb").read()) for image in os.listdir(output_dir)]

        elif file_type == 'images':
            images_info = [(file.name, file) for file in uploaded_files]

        # Check for any PDF files
        elif any(file.type == 'application/pdf' for file in uploaded_files):  
            images_info = []
            # Loop through each uploaded PDF
            for uploaded_file in uploaded_files: 
                pdf = pdfium.PdfDocument(uploaded_file)
                fn = uploaded_file.name
                for i in range(len(pdf)):
                    page = pdf[i]
                    image = page.render(scale=1.45).to_pil()
                    img_byte_arr = BytesIO()
                    image.save(img_byte_arr, format='JPEG')
                    if i == 0:
                        images_info.append((f"{fn.rsplit('.', 1)[0]}.jpg", img_byte_arr.getvalue()))
                    elif i > 0:
                        images_info.append((f"{fn.rsplit('.', 1)[0]}_page_{i + 1}.jpg", img_byte_arr.getvalue()))

if images_info:
    bg_image = None
    if add_bg:
        bg_file = st.file_uploader("Upload background image", type=["jpg", "jpeg", "png"])
        if bg_file:
            bg_image = resize_image(bg_file.read())
        else:
            # Use default background image if no new image is uploaded
            with open("./Bg.png", "rb") as file:
                default_bg_image = resize_image(file.read())
                bg_image = default_bg_image

    st.markdown("## Preview")
    #-------------------------------Now Download All Images button @ End of page-----------------------
    # if st.button("Download All Images", key="download_all"):
    #     zip_buffer = download_all_images_as_zip(images_info, remove_bg=remove_bg, add_bg=add_bg, bg_image=bg_image, resize_foreground=resize_fg, threshold=threshold,flip_horizontal=False, flip_vertical=False)
    #     st.download_button(
    #         label="Download All Images as ZIP",
    #         data=zip_buffer,
    #         file_name="all_images.zip",
    #         mime="application/zip"
    #     )
    #-------------------------------Now Download All Images button @ End of page-----------------------

    cols = st.columns(2)
    for i, (name, url_or_file) in enumerate(images_info):
        col = cols[i % 2]
        with col:
            if isinstance(url_or_file, str):
                url = convert_drive_link(url_or_file)
                image_content = download_image(url)
            elif isinstance(url_or_file, bytes):
                image_content = url_or_file
            else:
                image_content = url_or_file.read()

            if image_content:
                if remove_bg:
                    processed_image = remove_background(image_content)
                    ext = 'png'
                else:
                    size = (1290, 789) if "banner" in name.lower() else (1024, 1024)
                    processed_image = resize_image(image_content, size=size, aspect_ratio_threshold=threshold)
                    ext = "png"

                if add_bg and bg_image:
                    processed_image, dimensions = combine_with_background(processed_image, bg_image, resize_foreground=resize_fg)
                    ext = 'png'
                
                # Flip and rename options
                # Place flip options on the same row
                flip_col1, flip_col2,rename_col3 = st.columns(3)
                with flip_col1:
                    flip_horizontal = st.checkbox("Flip H🔁", key=f"flip_horizontal_{i}")
                with flip_col2:
                    flip_vertical = st.checkbox("Flip V🔃", key=f"flip_vertical_{i}")
                with rename_col3:
                    rename_image = st.checkbox(f"Rename", key=f"rename_checkbox_{i}")

                if rename_image:
                    new_name = st.text_input(f"Enter new name for {name} (without extension):", key=f"rename_input_{i}")
                    if new_name:
                        name = f"{new_name}.{ext}"  # Apply the new name

                # Apply flipping based on user selection
                if flip_horizontal or flip_vertical:
                    image = Image.open(BytesIO(processed_image))
                    processed_image = flip_image(image, flip_horizontal, flip_vertical)
                    img_byte_arr = BytesIO()
                    processed_image.save(img_byte_arr, format='PNG')
                    processed_image = img_byte_arr.getvalue()

                if processed_image:
                    # Store the processed image for downloading
                    images_info[i]=((name, processed_image))  # Add renamed/flipped image to the list
                    st.image(processed_image, caption=name)
                    st.download_button(
                        label=f"Download {name.rsplit('.', 1)[0]}",
                        data=processed_image,
                        file_name=f"{name.rsplit('.', 1)[0]}.{ext}",
                        mime=f"image/{ext}",
                        key=f"download_{i}"  # Unique key based on index
                    )
    #------------------------Download All Images Button-----------------------------------------
    if st.button("Download All Images", key="download_all"):
        zip_buffer = download_all_images_as_zip(images_info, remove_bg=remove_bg, add_bg=add_bg, bg_image=bg_image, resize_foreground=resize_fg, threshold=threshold,flip_horizontal=flip_horizontal, flip_vertical=flip_vertical)
        st.download_button(
            label="Download All Images as ZIP",
            data=zip_buffer,
            file_name="all_images.zip",
            mime="application/zip"
        )
 
